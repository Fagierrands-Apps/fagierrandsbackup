import os
import sys
import django
from datetime import date

sys.path.insert(0, '/home3/distinc3/fagiserver.fagtone.com/fagierrandsbackup')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fagierrandsbackup.settings')
django.setup()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import Count
from accounts.models import User
from orders.models import Order

# Date range
start = date(2026, 6, 15)
end   = date(2026, 7, 16)

# All July clients
clients = User.objects.filter(
    user_type__in=['user', 'client'],
).order_by('username')

# All orders in range
orders = (
    Order.objects
    .filter(client__in=clients, scheduled_date__gte=start, scheduled_date__lte=end)
    .values('client__username', 'scheduled_date')
    .annotate(count=Count('id'))
)

# Build lookup: {username: {date: count}}
data = {}
for row in orders:
    u = row['client__username']
    d = row['scheduled_date']
    data.setdefault(u, {})[d] = row['count']

# Date columns
from datetime import timedelta
days = [start + timedelta(days=i) for i in range((end - start).days + 1)]

# Build workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Errands Jun15-Jul16 2026"

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)
total_fill  = PatternFill("solid", fgColor="D9E1F2")

# Header row
ws.cell(1, 1, "Username").font = header_font
ws.cell(1, 1).fill = header_fill
ws.cell(1, 1).alignment = Alignment(horizontal='center')

for col, d in enumerate(days, 2):
    c = ws.cell(1, col, d.strftime("%d %b"))
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')

total_col = len(days) + 2
c = ws.cell(1, total_col, "TOTAL")
c.font = header_font
c.fill = header_fill
c.alignment = Alignment(horizontal='center')

# Data rows
for row_idx, client in enumerate(clients, 2):
    ws.cell(row_idx, 1, client.username)
    total = 0
    for col, d in enumerate(days, 2):
        count = data.get(client.username, {}).get(d, 0)
        if count:
            ws.cell(row_idx, col, count)
        total += count
    ws.cell(row_idx, total_col, total).fill = total_fill

# Totals row
total_row = clients.count() + 2
ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
for col in range(2, total_col + 1):
    col_total = sum(
        ws.cell(r, col).value or 0
        for r in range(2, total_row)
    )
    c = ws.cell(total_row, col, col_total)
    c.font = Font(bold=True)
    c.fill = total_fill

# Column widths
ws.column_dimensions['A'].width = 22
for col in range(2, total_col + 1):
    ws.column_dimensions[ws.cell(1, col).column_letter].width = 9

output = '/home3/distinc3/fagiserver.fagtone.com/errands_june15_to_july16.xlsx'
wb.save(output)
print(f"Saved: {output}")

# Summary print
print(f"\nTotal orders Jun 15–Jul 16: {sum(v for u in data.values() for v in u.values())}")
print(f"Clients with at least 1 order: {sum(1 for u in data if data[u])}")
