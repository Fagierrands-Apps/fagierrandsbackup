from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .models import Order
import logging

logger = logging.getLogger(__name__)


class ExportOrderView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            order = Order.objects.select_related(
                'client', 'assistant', 'handler', 'order_type',
                'pickup_location', 'delivery_location'
            ).get(pk=pk)
            
            # Check permissions
            if not (request.user == order.client or 
                    request.user == order.assistant or 
                    request.user == order.handler or 
                    request.user.user_type in ['admin', 'handler']):
                return Response(
                    {"error": "You don't have permission to export this order"},
                    status=status.HTTP_403_FORBIDDEN
                )

            wb = Workbook()
            ws = wb.active
            ws.title = f"Order {order.id}"

            # Header style
            header_font = Font(bold=True)
            
            # Order Information
            ws['A1'] = 'Order Information'
            ws['A1'].font = Font(bold=True, size=14)
            
            row = 3
            data = [
                ('Order ID', order.id),
                ('Title', order.title),
                ('Description', order.description),
                ('Order Type', order.order_type.name if order.order_type else ''),
                ('Status', order.get_status_display()),
                ('Client', order.client.get_full_name() or order.client.username),
                ('Client Phone', order.client.phone_number if hasattr(order.client, 'phone_number') else ''),
                ('Assistant', order.assistant.get_full_name() if order.assistant else 'Not Assigned'),
                ('Handler', order.handler.get_full_name() if order.handler else 'Not Assigned'),
                ('', ''),
                ('Pickup Address', order.pickup_address or ''),
                ('Delivery Address', order.delivery_address or ''),
                ('Recipient Name', order.recipient_name or ''),
                ('Contact Number', order.contact_number or ''),
                ('Alternative Contact', order.alternative_contact_name or ''),
                ('Alternative Number', order.alternative_contact_number or ''),
                ('', ''),
                ('Price (KSh)', f"{order.price:.2f}" if order.price else ''),
                ('Items Total (KSh)', f"{order.assistant_items_total:.2f}" if order.assistant_items_total else ''),
                ('Distance (km)', f"{order.distance:.2f}" if order.distance else ''),
                ('Estimated Duration (min)', order.estimated_duration or ''),
                ('Estimated Value (KSh)', f"{order.estimated_value:.2f}" if order.estimated_value else ''),
                ('', ''),
                ('Created At', order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else ''),
                ('Assigned At', order.assigned_at.strftime('%Y-%m-%d %H:%M:%S') if order.assigned_at else ''),
                ('Started At', order.started_at.strftime('%Y-%m-%d %H:%M:%S') if order.started_at else ''),
                ('Completed At', order.completed_at.strftime('%Y-%m-%d %H:%M:%S') if order.completed_at else ''),
                ('Cancelled At', order.cancelled_at.strftime('%Y-%m-%d %H:%M:%S') if order.cancelled_at else ''),
            ]
            
            for label, value in data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'] = value
                row += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 50

            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="order_{order.id}.xlsx"'
            wb.save(response)
            
            return response

        except Order.DoesNotExist:
            return Response(
                {"error": "Order not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error exporting order {pk}: {str(e)}")
            return Response(
                {"error": "Failed to export order"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
