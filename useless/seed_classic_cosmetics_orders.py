"""
Seed Classic Cosmetics orders into the database.
Run: python manage.py shell < seed_classic_cosmetics_orders.py
"""

import os
import django
from datetime import datetime
from decimal import Decimal
import re

if __name__ == "__main__":
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fagierrandsbackup.settings')
    django.setup()

from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from orders.models import Order, OrderType

User = get_user_model()

# Load Excel data
wb = load_workbook('/home/jarvis/Downloads/seed orders.xlsx')
ws = wb.active
all_rows = list(ws.iter_rows(values_only=True))

headers = all_rows[2]
client_name_idx = headers.index('Client name')
phone_idx = headers.index('Phone number')
date_idx = headers.index('Date')
errand_status_idx = headers.index('errand status')
pickup_location_idx = headers.index('Pick up -location')
dropoff_location_idx = headers.index('Drop off -location')
estimated_value_idx = headers.index('estimated value   of the product')
distance_idx = headers.index('delivery distance(km)')
load_category_idx = headers.index('load category')
payment_status_idx = headers.index('Payment status')

# Get or create Classic Cosmetics user
classic_user, created = User.objects.get_or_create(
    username='classiccosmetics',
    defaults={
        'email': 'classiccosmetics@gmail.com',
        'first_name': 'Classic',
        'last_name': 'Cosmetic User',
        'is_active': True
    }
)
if created:
    classic_user.set_password('temppassword123')
    classic_user.save()
    print(f"Created user: {classic_user.username}")
else:
    print(f"User already exists: {classic_user.username}")

# Map status and order type
status_map = {
    'pending': 'pending',
    'assigned': 'assigned',
    'in progress': 'in_progress',
    'in_progress': 'in_progress',
    'payment pending': 'payment_pending',
    'payment_pending': 'payment_pending',
    'completed': 'completed',
    'cancelled': 'cancelled',
}

category_to_order_type = {
    'shopping': 'Shopping',
    'delivery': 'Delivery',
    'cargo': 'Cargo Delivery',
    'banking': 'Banking',
    'handyman': 'Handyman',
}

def parse_distance(distance_str):
    """Extract numeric distance from string like '5.1km'"""
    if not distance_str:
        return None
    match = re.search(r'(\d+\.?\d*)', str(distance_str))
    return float(match.group(1)) if match else None

def parse_value(value_str):
    """Extract numeric value from string like '5000ksh'"""
    if not value_str:
        return None
    match = re.search(r'(\d+)', str(value_str).replace(',', ''))
    return Decimal(match.group(1)) if match else None

def parse_date(date_val):
    """Parse date from Excel"""
    if not date_val:
        return None
    if isinstance(date_val, str):
        try:
            return datetime.strptime(date_val, '%d.%m.%Y').date()
        except:
            return None
    return date_val

# Classic Cosmetics phone numbers
classic_phones = ['721420878', '722995300']

# Extract and seed orders
orders_created = 0
orders_skipped = 0

for row in all_rows[3:]:
    if not row[client_name_idx] or not row[phone_idx]:
        continue
    
    phone = str(int(row[phone_idx])) if isinstance(row[phone_idx], (int, float)) else str(row[phone_idx]).strip()
    
    if phone not in classic_phones:
        continue
    
    try:
        # Parse fields
        title = str(row[client_name_idx]).strip()
        description = f"Order from {row[pickup_location_idx]} to {row[dropoff_location_idx]}"
        
        # Get order type
        category = str(row[load_category_idx]).lower().strip() if row[load_category_idx] else 'delivery'
        order_type_name = category_to_order_type.get(category, 'Delivery')
        order_type = OrderType.objects.get(name=order_type_name)
        
        # Get status
        status_raw = str(row[errand_status_idx]).lower().strip() if row[errand_status_idx] else 'pending'
        status = status_map.get(status_raw, 'pending')
        
        # Parse numeric fields
        distance = parse_distance(row[distance_idx])
        estimated_value = parse_value(row[estimated_value_idx])
        
        # Create order
        order = Order.objects.create(
            client=classic_user,
            order_type=order_type,
            title=title[:255],
            description=description[:500],
            pickup_address=str(row[pickup_location_idx])[:255] if row[pickup_location_idx] else '',
            delivery_address=str(row[dropoff_location_idx])[:255] if row[dropoff_location_idx] else '',
            contact_number=phone,
            distance=distance,
            estimated_value=estimated_value,
            status=status,
            price_finalized=False
        )
        
        # Calculate price if distance available
        if distance:
            order.price = order.calculate_price()
            order.save()
        
        orders_created += 1
        print(f"✓ Created order: {title} ({status})")
        
    except Exception as e:
        orders_skipped += 1
        print(f"✗ Skipped: {row[client_name_idx]} - {str(e)}")

print(f"\n{'='*60}")
print(f"SEEDING COMPLETE")
print(f"Orders created: {orders_created}")
print(f"Orders skipped: {orders_skipped}")
print(f"Total: {orders_created + orders_skipped}")
